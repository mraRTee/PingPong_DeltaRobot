################################################################################## ez itt a jó!!!!!!!!!!!!!!
import cv2
import numpy as np
import matplotlib.pyplot as plt
import serial
from sklearn.metrics import r2_score
import math
from datetime import datetime
from openpyxl import Workbook

book = Workbook()
sheet = book.active
row=1
column=1
# ser = serial.Serial(
#         port='COM3',
#         baudrate=115200,
#         parity=serial.PARITY_NONE,
#         stopbits=serial.STOPBITS_ONE,
#         bytesize=serial.EIGHTBITS,
#         timeout=1
#     )


listy_felso = []
listx_felso = []
listy_also = []
listx_also = []
list_kiirx_f = []
list_kiiry_f = []
list_kiirx_l = []
list_kiiry_l = []
seged=0
atlag=0

x_a = 0
y_a = 0
x_b = 0
y_b = 0
kd1 = ""
kd2 = ""
s_1 = False
s_2 = False
cn_1 = 0
cn_2 = 0
cim1 = ""
txt = ".txt"
e = 0
x_kozepe_felso = 367 # ezt majd be kell állítani mindig, egy következő videónál!!!!!
                # azért kell, hogy beállítsuk, hogy pozitív vagy negatív Y értéke lesz a NEUTRÁL ponthoz képest
szorzo_felso = 0.17   # ezt is!!!!!
                # ez ahhoz kell, hogy megadjuk 1 pixel hány mm a valóságban
# a regressio_felso függvényben úgy néz ki,
# hogy: real_x=(int((int((koef1*robot_helyzete_x_tengely*robot_helyzete_x_tengely)+(koef2*robot_helyzete_x_tengely)+koef3)-x_kozepe)*szorzo))
y_kozepe_also = -550 # azért Y, mert a reális tengelyek szempontjából az Y értékekről van szó
szorzo_also = 0.13 # be kell állítani a lenti kamera kalibrációja alapján
z_kozepe_also=391
origo_z=-93 # nullától az asztal távolsága
origo_x=556
origo_y=367

#mm
effector_radius = 24.0        #effector radius
base_radius = 100.0        #base radius
alkar_hossz = 800.0      #alkar
felkar_hossz = 300.0       #felkar

# Trigonometric constants
s_allando = 165 * 2
sqrt3 = math.sqrt(3.0)
pi = 3.141592653
sin120 = sqrt3 / 2.0
cos120 = -0.5
tan60 = sqrt3
sin30 = 0.5
tan30 = 1.0 / sqrt3

def angle_yz(x0, y0, z0, theta=None):
    y1 = -0.5 * 0.57735 * base_radius  # f/2 * tg 30
    y0 -= 0.5 * 0.57735 * effector_radius  # shift center to edge
    # z = a + b*y
    a = (x0 * x0 + y0 * y0 + z0 * z0 + felkar_hossz * felkar_hossz - alkar_hossz * alkar_hossz - y1 * y1) / (2.0 * z0)
    b = (y1 - y0) / z0

    # discriminant
    d = -(a + b * y1) * (a + b * y1) + felkar_hossz * (b * b * felkar_hossz + felkar_hossz)
    if d < 0:
        return [1, 0]  # non-existing povar.  return error, theta

    yj = (y1 - a * b - math.sqrt(d)) / (b * b + 1)  # choosing outer povar
    zj = a + b * yj
    theta = math.atan(-zj / (y1 - yj)) * 180.0 / pi + (180.0 if yj > y1 else 0.0)

    return [0, theta]  # return error, theta

def inverse(x0, y0, z0):
    global theta1
    global theta2
    global theta3
    theta1 = 0
    theta2 = 0
    theta3 = 0
    status = angle_yz(x0, y0, z0)

    if status[0] == 0:
        theta1 = status[1]
        status = angle_yz(x0 * cos120 + y0 * sin120,
                          y0 * cos120 - x0 * sin120,
                          z0,
                          theta2)
    if status[0] == 0:
        theta2 = status[1]
        status = angle_yz(x0 * cos120 - y0 * sin120,
                          y0 * cos120 + x0 * sin120,
                          z0,
                          theta3)
    theta3 = status[1]

    return [status[0], theta1, theta2, theta3]

def kuldes(atad):

    atad=str(atad)
    ser.write(atad.encode())

def regress_felso(listx_felso,listy_felso):
    z = np.polyfit(listx_felso,listy_felso, 2)
    egyenlet_1 = np.poly1d(z)
    egyenlet_1 = str(egyenlet_1)
    # print(egyenlet_1)
    robot_helyzete_x_tengely = 1  # előzetes kalkuláció!!! - HOL VAN A ROBOT NULL PONTJA AZ Y TENGELYEN?????? pixelben
    a=egyenlet_1.split("\n")
    b=a[1]
    b=b.split(" ")
    seged1=0
    seged2=""
    # print(b)
    koef1 = float(b[0])
    koef2 = float(b[3])
    koef3 = float(b[6])
    for neg in b:
        if seged1==0 and neg == '-':
            seged2 = ("-"+b[seged1 + 1])
            koef1 = float(seged2)
        if seged1>=2 and seged1<=4 and neg == '-':
            seged2 = ("-"+b[seged1+1])
            koef2 = float(seged2)
        if seged1 >= 5 and neg == '-':
            seged2 = ("-"+b[seged1 + 1])
            koef3 = float(seged2)
        seged1 += 1
    # átszámítjuk valódi koordinátára és kiszámítjuk a robot null helyzetéhez viszonyított, várható érkezés
    global real_y_1
    real_y_1=((x_kozepe_felso-((koef1 * robot_helyzete_x_tengely * robot_helyzete_x_tengely) + (koef2 * robot_helyzete_x_tengely) + koef3)) * szorzo_felso)
    #####################################################################################################
    atkuld_ertek = inverse(0, real_y_1*10, -732)
    if (round(atkuld_ertek[1])) == 0:
        send_string_1 = "0000"
    if (round(atkuld_ertek[2])) == 0:
        send_string_2 = "0000"
    if (round(atkuld_ertek[3])) == 0:
        send_string_3 = "0000"
    if (round(atkuld_ertek[1])) <= -100:
        send_string_1="1"+str(-round(atkuld_ertek[1]))
    if (round(atkuld_ertek[2])) <= -100:
        send_string_2="1"+str(-round(atkuld_ertek[2]))
    if (round(atkuld_ertek[3])) <= -100:
        send_string_3="1"+str(-round(atkuld_ertek[3]))
    if (round(atkuld_ertek[1])) <= -10 and (round(atkuld_ertek[1])) > -100:
        send_string_1="10"+str(-round(atkuld_ertek[1]))
    if (round(atkuld_ertek[2])) <= -10 and (round(atkuld_ertek[2])) > -100:
        send_string_2="10"+str(-round(atkuld_ertek[2]))
    if (round(atkuld_ertek[3])) <= -10 and (round(atkuld_ertek[3])) > -100:
        send_string_3="10"+str(-round(atkuld_ertek[3]))
    if (round(atkuld_ertek[1])) < 0 and (round(atkuld_ertek[1])) > -10:
        send_string_1="100"+str(-round(atkuld_ertek[1]))
    if (round(atkuld_ertek[2])) < 0 and (round(atkuld_ertek[2])) > -10:
        send_string_2="100"+str(-round(atkuld_ertek[2]))
    if (round(atkuld_ertek[3])) < 0 and (round(atkuld_ertek[3])) > -10:
        send_string_3="100"+str(-round(atkuld_ertek[3]))
    if (round(atkuld_ertek[1])) < 10 and (round(atkuld_ertek[1])) > 0:
        send_string_1="000"+str(round(atkuld_ertek[1]))
    if (round(atkuld_ertek[2])) < 10 and (round(atkuld_ertek[2])) > 0:
        send_string_2="000"+str(round(atkuld_ertek[2]))
    if (round(atkuld_ertek[3])) < 10 and (round(atkuld_ertek[3])) > 0:
        send_string_3="000"+str(round(atkuld_ertek[3]))
    if ((round(atkuld_ertek[1])) >= 10) and ((round(atkuld_ertek[1])) <100):
        send_string_1="00"+str(round(atkuld_ertek[1]))
    if ((round(atkuld_ertek[2])) >= 10) and ((round(atkuld_ertek[2])) <100):
        send_string_2="00"+str(round(atkuld_ertek[2]))
    if ((round(atkuld_ertek[3])) >= 10) and ((round(atkuld_ertek[1])) <100):
        send_string_3="00"+str(round(atkuld_ertek[3]))
    if ((round(atkuld_ertek[1])) >= 100):
        send_string_1="0"+str(round(atkuld_ertek[1]))
    if ((round(atkuld_ertek[2])) >= 100):
        send_string_2="0"+str(round(atkuld_ertek[2]))
    if ((round(atkuld_ertek[3])) >= 100):
        send_string_3="0"+str(round(atkuld_ertek[3]))

    print("Fenti adat 1:", send_string_1)
    print("Fenti adat 2:",send_string_2)
    print("Fenti adat 3:",send_string_3)
    # uj=datetime.now()
    # uj=str(uj)
    # uj = uj[17:25]
    # print(uj)
    # row=2
    # sheet.cell(row, column).value = uj
    # book.save('idomeres.xlsx')
    # kuldes(send_string_1+send_string_2+send_string_3)

def real_koordinatak(z_also,x_also,x_felso,y_felso):
    real_z=origo_z+((z_kozepe_also-z_also)*szorzo_also)
    real_x=(x_also-origo_x)*szorzo_also
    real_y=(origo_y-y_felso)*szorzo_felso
    atkuld_ertek=inverse(real_x*10,real_y*10,real_z*10)
    if (round(atkuld_ertek[1])) == 0:
        send_string_1 = "0000"
    if (round(atkuld_ertek[2])) == 0:
        send_string_2 = "0000"
    if (round(atkuld_ertek[3])) == 0:
        send_string_3 = "0000"
    if (round(atkuld_ertek[1])) <= -100:
        send_string_1="1"+str(-round(atkuld_ertek[1]))
    if (round(atkuld_ertek[2])) <= -100:
        send_string_2="1"+str(-round(atkuld_ertek[2]))
    if (round(atkuld_ertek[3])) <= -100:
        send_string_3="1"+str(-round(atkuld_ertek[3]))
    if (round(atkuld_ertek[1])) <= -10 and (round(atkuld_ertek[1])) > -100:
        send_string_1="10"+str(-round(atkuld_ertek[1]))
    if (round(atkuld_ertek[2])) <= -10 and (round(atkuld_ertek[2])) > -100:
        send_string_2="10"+str(-round(atkuld_ertek[2]))
    if (round(atkuld_ertek[3])) <= -10 and (round(atkuld_ertek[3])) > -100:
        send_string_3="10"+str(-round(atkuld_ertek[3]))
    if (round(atkuld_ertek[1])) < 0 and (round(atkuld_ertek[1])) > -10:
        send_string_1="100"+str(-round(atkuld_ertek[1]))
    if (round(atkuld_ertek[2])) < 0 and (round(atkuld_ertek[2])) > -10:
        send_string_2="100"+str(-round(atkuld_ertek[2]))
    if (round(atkuld_ertek[3])) < 0 and (round(atkuld_ertek[3])) > -10:
        send_string_3="100"+str(-round(atkuld_ertek[3]))
    if (round(atkuld_ertek[1])) < 10 and (round(atkuld_ertek[1])) > 0:
        send_string_1="000"+str(round(atkuld_ertek[1]))
    if (round(atkuld_ertek[2])) < 10 and (round(atkuld_ertek[2])) > 0:
        send_string_2="000"+str(round(atkuld_ertek[2]))
    if (round(atkuld_ertek[3])) < 10 and (round(atkuld_ertek[3])) > 0:
        send_string_3="000"+str(round(atkuld_ertek[3]))
    if ((round(atkuld_ertek[1])) >= 10) and ((round(atkuld_ertek[1])) <100):
        send_string_1="00"+str(round(atkuld_ertek[1]))
    if ((round(atkuld_ertek[2])) >= 10) and ((round(atkuld_ertek[2])) <100):
        send_string_2="00"+str(round(atkuld_ertek[2]))
    if ((round(atkuld_ertek[3])) >= 10) and ((round(atkuld_ertek[1])) <100):
        send_string_3="00"+str(round(atkuld_ertek[3]))
    if ((round(atkuld_ertek[1])) >= 100):
        send_string_1="0"+str(round(atkuld_ertek[1]))
    if ((round(atkuld_ertek[2])) >= 100):
        send_string_2="0"+str(round(atkuld_ertek[2]))
    if ((round(atkuld_ertek[3])) >= 100):
        send_string_3="0"+str(round(atkuld_ertek[3]))

    print("Lenti adat 1:", send_string_1)
    print("Lenti adat 2:", send_string_2)
    print("Lenti adat 3:", send_string_3)
    # uj = datetime.now()
    # uj = str(uj)
    # uj=uj[17:25]
    # print(uj)
    # global column
    # row=3
    # sheet.cell(row, column).value = uj
    # column+=1
    # book.save('idomeres.xlsx')
    print("\n")
    # kuldes(send_string_1+send_string_2+send_string_3)
    # print(uj)

cap1 = cv2.VideoCapture("17_f.mp4")
cap2 = cv2.VideoCapture("17_l.mp4")
# cap1 = cv2.VideoCapture(0)
# cap2 = cv2.VideoCapture(1)

'''def kiir():
    cim1 = str(e) + "y_felso" + txt
    with open(cim1, 'w+') as f:
        for item in listy_felso:
            f.write("%s\n" % item)
    cim1 = str(e) + "x_felso" + txt
    with open(cim1, 'w+') as f:
        for item in listx_felso:
            f.write("%s\n" % item)
    cim1 = str(e) + "y_also" + txt
    with open(cim1, 'w+') as f:
        for item in listy_also:
            f.write("%s\n" % item)
    cim1 = str(e) + "x_also" + txt
    with open(cim1, 'w+') as f:
        for item in listx_also:
            f.write("%s\n" % item)'''


def plot():
    plt.plot(listx_felso, listy_felso, 'ro')
    plt.axis([0, 1280, 720, 0])
    plt.plot(listx_also, listy_also, 'go')
    plt.axis([0, 1280, 720, 0])
    plt.show()


while 1:  # folyamatosan veszi a jelet
    #időmérés
    global ido
    ido=datetime.now()


    _, vid1 = cap1.read()  # minden kockát olvas
    _, vid2 = cap2.read()  # minden kockát olvas
    try:
        hsv1 = cv2.cvtColor(vid1, cv2.COLOR_BGR2HSV)  # ez itt a konverzió
        hsv2 = cv2.cvtColor(vid2, cv2.COLOR_BGR2HSV)  # ez itt a konverzió
    except:
        '''with open("y_11_felső_full.txt", 'w+') as f:
            for item in list_kiirx_f:
                f.write("%s\n" % item)
        with open("x_11_felső_full.txt", 'w+') as f:
            for item in list_kiiry_f:
                f.write("%s\n" % item)
        with open("y_11_alsó_full.txt", 'w+') as f:
            for item in list_kiirx_l:
                f.write("%s\n" % item)
        with open("x_11_alsó_full.txt", 'w+') as f:
            for item in list_kiiry_l:
                f.write("%s\n" % item)'''
        plot()
    # lower_orange1 = np.array([30, 130, 130])
    # upper_orange1 = np.array([85, 255, 255])  # új videóknál ez az atom!!!
    # lower_orange2 = np.array([30, 130, 130])
    # upper_orange2 = np.array([85, 255, 255])

    lower_orange1 = np.array([1, 170, 200])             # narancs jó
    upper_orange1 = np.array([18, 255, 255])
    lower_orange2 = np.array([1, 180, 180])             #narancs jó
    upper_orange2 = np.array([18, 255, 255])

    # lower_orange2 = np.array([30, 130, 130]) #zöld jó
    # upper_orange2 = np.array([85, 255, 255])

    # lower_orange2 = np.array([120, 100, 100]) # lila
    # upper_orange2 = np.array([165, 150, 255])

    mask1 = cv2.inRange(hsv1, lower_orange1, upper_orange1)  # ez maszkolja ki
    mask2 = cv2.inRange(hsv2, lower_orange2, upper_orange2)  # ez maszkolja ki
    # res1 = cv2.bitwise_and(vid1, vid1, mask=mask1)  # a kimaszkolt és igazi kép összeÉSelve.

    # external_poly_1_a = np.array([[[0, 0], [150, 0], [110, 720], [0, 720]]], dtype=np.int32)
    # external_poly_2_a = np.array([[[0, 0], [1280, 0], [1280, 50], [0, 15]]], dtype=np.int32)
    # external_poly_3_a = np.array([[[1040, 0], [1280, 0], [1280, 720], [1030, 720]]], dtype=np.int32)
    # external_poly_4_a = np.array([[[0, 660], [1280, 680], [1280, 720], [0, 720]]], dtype=np.int32)
    #
    # external_poly_1_b = np.array([[[820,80], [820, 130], [760, 130], [760, 80]]], dtype=np.int32)
    # external_poly_2_b = np.array([[[0, 0], [1280, 0], [1280, 250], [0, 250]]], dtype=np.int32)
    # external_poly_3_b = np.array([[[1100, 0], [1280, 0], [1280, 720], [1100, 720]]], dtype=np.int32)
    # external_poly_4_b = np.array([[[0, 650], [1280, 650], [1280, 720], [0, 720]]], dtype=np.int32)
    # [[[10, 10], [100, 10], [100, 100], [10, 100]]] kis négyzet
    # https: // stackoverflow.com / questions / 56813343 / masking - out - a - specific - region - in -opencv - python

    # ez a rész a konkrét videó maszkolás rész, akkor is kell ha nem kell látnunk a videót #####
    # cv2.fillPoly(mask1, external_poly_1_a, (0, 0, 0))
    # cv2.fillPoly(mask1, external_poly_2_a, (0, 0, 0))
    # cv2.fillPoly(mask1, external_poly_3_a, (0, 0, 0))
    # cv2.fillPoly(mask1, external_poly_4_a, (0, 0, 0))

    # cv2.fillPoly(mask2, external_poly_1_b, (0, 0, 0))
    # cv2.fillPoly(mask2, external_poly_2_b, (0, 0, 0))
    # cv2.fillPoly(mask2, external_poly_3_b, (0, 0, 0))
    # cv2.fillPoly(mask2, external_poly_4_b, (0, 0, 0))
    ############################################################################################

    # ez a rész csak a videós megjelenítéshez kell #############################################
    # cv2.fillPoly(vid1, external_poly_1_a, (0, 0, 0))
    # cv2.fillPoly(vid1, external_poly_2_a, (0, 0, 0))
    # cv2.fillPoly(vid1, external_poly_3_a, (0, 0, 0))
    # cv2.fillPoly(vid1, external_poly_4_a, (0, 0, 0))

    # cv2.fillPoly(vid2, external_poly_1_b, (0, 0, 0))
    # cv2.fillPoly(vid2, external_poly_2_b, (0, 0, 0))
    # cv2.fillPoly(vid2, external_poly_3_b, (0, 0, 0))
    # cv2.fillPoly(vid2, external_poly_4_b, (0, 0, 0))
    ############################################################################################

    cv2.line(vid1, (int(y_a), int(x_a)), (int(y_a), int(x_a)), (0, 0, 255), 15)
    cv2.line(vid2, (int(y_b), int(x_b)), (int(y_b), int(x_b)), (255, 0, 0), 15)

    x_a = 0
    y_a = 0
    x_b = 0
    y_b = 0
    cv2.imshow('vid1', vid1)  # mit mutat
    cv2.imshow('vid2', vid2)
    cv2.waitKey(1)
    # if cv2.waitKey(1)&0xff==0:
    #     break# k = cv2.waitKey(1) & 0xFF #itt ez nem tudom mi volt, de ugyanúgy működik így

    npImg1 = np.asarray(mask1)  # melyikről vegye a mintát
    npImg2 = np.asarray(mask2)
    kd_b = np.argwhere(npImg2 == 255)
    kd_a = np.argwhere(npImg1 == 255)
    kd_a = str(kd_a)
    kd_b = str(kd_b)
    for i in range(15):
        # ha nincs külön minden if-re egy try akkor csak akkor fog működni, ha a fenti kamera detektál,
        # lehet az még hasznos lehet

        try:
            if kd_a[i] == ' ' and kd1 and not x_a:
                x_a = kd1
                x_a = int(x_a)
                kd1 = ""
        except:
            pass
        try:
            if kd_a[i] == ' ' and kd1 and x_a:
                y_a = kd1
                y_a = int(y_a)
                kd1 = ""
        except:
            pass
        try:
            if kd_a[i] == '0' or kd_a[i] == '1' or kd_a[i] == '2' or kd_a[i] == '3' or kd_a[i] == '4' or kd_a[i] == '5' \
                    or kd_a[i] == '6' or kd_a[i] == '7' or kd_a[i] == '8' or kd_a[i] == '9':
                kd1 += kd_a[i]
        except:
            pass
        try:
            if kd_b[i] == ' ' and kd2 and not x_b:
                x_b = kd2
                x_b = int(x_b)
                kd2 = ""
        except:
            pass
        try:
            if kd_b[i] == ' ' and kd2 and x_b:
                y_b = kd2
                y_b = int(y_b)
                kd2 = ""
        except:
            pass
        try:
            if kd_b[i] == '0' or kd_b[i] == '1' or kd_b[i] == '2' or kd_b[i] == '3' or kd_b[i] == '4' or kd_b[i] == '5' \
                    or kd_b[i] == '6' or kd_b[i] == '7' or kd_b[i] == '8' or kd_b[i] == '9':
                kd2 += kd_b[i]
        except:
            pass
    kd1 = ""
    kd2 = ""

    if x_a != 0 and y_a != 0:
        listy_felso.append(int(x_a))
        listx_felso.append(int(y_a))
        if x_b != 0 and y_b != 0:
            listy_also.append(int(x_b))
            listx_also.append(int(y_b))
        if cn_1 == 15:
            # plot()
            # print(listx_felso)
            # print(listy_felso)
            # ido=str(ido)
            # ido=ido[17:25]
            # print(ido)
            # row = 1
            # sheet.cell(row, column).value = ido
            # book.save('idomeres.xlsx')
            regress_felso(listx_felso,listy_felso)
        if cn_1 > 0 and listx_felso[cn_1] > listx_felso[cn_1-1]:
            # plot()
            cn_1 = -1
            listy_felso = []  # kinullázza az aktuális listát
            listx_felso = []
        if cn_1 != 0 and x_a == 0 and y_a == 0:
            listy_felso = []  # kinullázza az aktuális listát
            listx_felso = []
            cn_1 = 0
        if cn_1 > 0 and len(listy_also) > 1:
            if listy_also[len(listy_also)-1] < listy_also[len(listy_also)-2] and listx_also[len(listy_also)-1] < 1100:
                if seged==0:
                    seged=cn_1
                if cn_1-seged==5:
                    real_koordinatak(listy_also[-1], listx_also[-1], listx_felso[-1], listy_felso[-1])
                    # plot()
            if listy_also[len(listy_also)-1] > listy_also[len(listy_also)-2]:
                listy_also = []  # kinullázza az aktuális listát
                listx_also = []
                seged=0
        cn_1 += 1

cv2.destroyAllWindows()



